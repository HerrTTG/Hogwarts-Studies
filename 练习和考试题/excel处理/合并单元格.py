import openpyxl

wb=openpyxl.Workbook()
sheet=wb.active
sheet['A1'].value='Hello world'
sheet['C5'].value='Help'

sheet.merge_cells('A1:D3')
sheet.merge_cells('C5:D5')


wb.save('./合并单元格.xlsx')