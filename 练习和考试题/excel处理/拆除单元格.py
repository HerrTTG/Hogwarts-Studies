import openpyxl

wb=openpyxl.load_workbook('./合并单元格.xlsx')
sheet=wb.active
sheet.unmerge_cells('A1:D3')
sheet.unmerge_cells('C5:D5')

wb.save('./合并后拆除单元格.xlsx')