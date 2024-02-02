import openpyxl

wb=openpyxl.load_workbook('.//example.xlsx')

print(wb.sheetnames)

sheet=wb['Sheet1']

print(sheet.print_titles)

print(wb.active)

c=sheet['B1']

print(c.coordinate,c.value,c.row,c.column)

c=sheet.cell(row=1,column=3)
print(c.coordinate,c.value,c.row,c.column)