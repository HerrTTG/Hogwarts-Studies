import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string

wb=openpyxl.load_workbook('.//example.xlsx')

sheet1=wb['Sheet1']

print(sheet1.max_column)
#默认情况下列展示为index
print(get_column_letter(sheet1.max_column))
#转换需要用到get_column_letter函数
