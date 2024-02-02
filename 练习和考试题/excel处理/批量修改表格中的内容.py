#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('./produceSales.xlsx')
sheet = wb['Sheet']

#要修改的产品和其修改值生成一个字典
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row+1): # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    #sheet['A'+str(rowNum)]
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
        #sheet['B'+str(rowNum)]

wb.save('./updatedProduceSales.xlsx')